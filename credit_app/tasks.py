import os
import logging
from datetime import datetime
from django.db import connection
from celery import shared_task
from django.conf import settings

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def ingest_customer_data(self, file_path=None):
    """Background task to ingest customer data from Excel file."""
    try:
        import openpyxl
        from credit_app.models import Customer

        if file_path is None:
            file_path = os.path.join(settings.DATA_DIR, 'customer_data.xlsx')

        if not os.path.exists(file_path):
            logger.warning(f"Customer data file not found: {file_path}")
            return {'status': 'skipped', 'reason': 'file not found'}

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        customers_created = 0
        customers_updated = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = dict(zip(headers, row))

            customer_id = data.get('customer_id') or data.get('Customer ID')
            first_name = data.get('first_name') or data.get('First Name', '')
            last_name = data.get('last_name') or data.get('Last Name', '')
            phone_number = data.get('phone_number') or data.get('Phone Number', 0)
            monthly_salary = data.get('monthly_salary') or data.get('Monthly Salary', 0)
            approved_limit = data.get('approved_limit') or data.get('Approved Limit', 0)
            current_debt = data.get('current_debt') or data.get('Current Debt', 0.0)

            obj, created = Customer.objects.update_or_create(
                customer_id=int(customer_id),
                defaults={
                    'first_name': str(first_name),
                    'last_name': str(last_name),
                    'phone_number': int(phone_number) if phone_number else 0,
                    'monthly_salary': int(monthly_salary) if monthly_salary else 0,
                    'approved_limit': int(approved_limit) if approved_limit else 0,
                    'current_debt': float(current_debt) if current_debt else 0.0,
                }
            )
            if created:
                customers_created += 1
            else:
                customers_updated += 1

        with connection.cursor() as cursor:
            cursor.execute("SELECT setval('customers_customer_id_seq', (SELECT MAX(customer_id) FROM customers));")

        logger.info(f"Customer ingestion done: {customers_created} created, {customers_updated} updated")
        return {'status': 'success', 'created': customers_created, 'updated': customers_updated}

    except Exception as exc:
        logger.error(f"Error ingesting customer data: {exc}")
        raise self.retry(exc=exc, countdown=5)


@shared_task(bind=True, max_retries=3)
def ingest_loan_data(self, file_path=None):
    """Background task to ingest loan data from Excel file."""
    try:
        import openpyxl
        from credit_app.models import Customer, Loan

        if file_path is None:
            file_path = os.path.join(settings.DATA_DIR, 'loan_data.xlsx')

        if not os.path.exists(file_path):
            logger.warning(f"Loan data file not found: {file_path}")
            return {'status': 'skipped', 'reason': 'file not found'}

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        loans_created = 0
        loans_skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = dict(zip(headers, row))

            customer_id = data.get('customer_id') or data.get('Customer ID')
            loan_id = data.get('loan_id') or data.get('Loan ID')
            loan_amount = data.get('loan_amount') or data.get('Loan Amount', 0)
            tenure = data.get('tenure') or data.get('Tenure', 0)
            interest_rate = data.get('interest_rate') or data.get('Interest Rate', 0)
            monthly_repayment = data.get('monthly_repayment') or data.get('Monthly Repayment', 0)
            emis_paid_on_time = data.get('EMIs paid on Time') or data.get('emis_paid_on_time', 0)
            start_date = data.get('start_date') or data.get('Date of Approval')
            end_date = data.get('end_date') or data.get('End Date')

            try:
                customer = Customer.objects.get(customer_id=int(customer_id))
            except Customer.DoesNotExist:
                loans_skipped += 1
                continue
            with connection.cursor() as cursor:
                cursor.execute("SELECT setval('loans_loan_id_seq', (SELECT MAX(loan_id) FROM loans));")
            # Parse dates
            def parse_date(d):
                if d is None:
                    return None
                if isinstance(d, datetime):
                    return d.date()
                if hasattr(d, 'date'):
                    return d.date()
                try:
                    return datetime.strptime(str(d), '%Y-%m-%d').date()
                except Exception:
                    return None

            Loan.objects.update_or_create(
                loan_id=int(loan_id),
                defaults={
                    'customer': customer,
                    'loan_amount': float(loan_amount) if loan_amount else 0.0,
                    'tenure': int(tenure) if tenure else 0,
                    'interest_rate': float(interest_rate) if interest_rate else 0.0,
                    'monthly_repayment': float(monthly_repayment) if monthly_repayment else 0.0,
                    'emis_paid_on_time': int(emis_paid_on_time) if emis_paid_on_time else 0,
                    'start_date': parse_date(start_date),
                    'end_date': parse_date(end_date),
                }
            )
            loans_created += 1

        logger.info(f"Loan ingestion done: {loans_created} processed, {loans_skipped} skipped")
        return {'status': 'success', 'processed': loans_created, 'skipped': loans_skipped}

    except Exception as exc:
        logger.error(f"Error ingesting loan data: {exc}")
        raise self.retry(exc=exc, countdown=5)
